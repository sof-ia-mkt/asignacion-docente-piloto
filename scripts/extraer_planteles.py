#!/usr/bin/env python3
"""
Extrae y limpia los datos de los 4 planteles de CENYCA desde el Excel de proyección.
Salida: db/seed_data/proyeccion.json (consumido por cargar_planteles.mjs).

Migración FIEL: toda fila con ID + materia entra como slot, aunque le falte
el grupo (PALMAS tiene ~180 filas sin grupo capturado). En esos casos grupo=null
y se le da sentido en la fase 2 (NO se inventa con forward-fill: uniría grupos distintos).

Decisiones de limpieza (heredadas de extraer_casablanca.py):
- Plantel se deriva del NOMBRE de la hoja (la columna PLANTEL viene vacía en CB).
- Llave de slot: ID + plantel (el ID NO es único entre planteles).
- Fechas: parser tolerante para formatos en texto libre.
- Columna de grupo detectada por patrón PLAN_Gnn (header a veces vacío en CB).
- Mayo = historial. Septiembre = misma estructura, docente vacío.
"""
import openpyxl, re, json, datetime, unicodedata
from collections import Counter, defaultdict
from pathlib import Path

EXCEL = Path.home() / "Downloads" / "PROYECCIÓN MAYO - AGOSTO -ACTUALIZADA.xlsx"
OUT = Path(__file__).resolve().parent.parent / "db" / "seed_data" / "proyeccion.json"

# Hoja del Excel -> (nombre de plantel canónico, sufijo en clave de grupo)
PLANTELES = {
    "CASA BLANCA": ("CASA BLANCA", "CB"),
    "OTAY": ("OTAY", "OT"),
    "TECATE": ("TECATE", "TC"),
    "PALMAS": ("PALMAS", "PL"),
}

MESES = {
    "ENERO": 1, "FEBRERO": 2, "MARZO": 3, "ABRIL": 4, "MAYO": 5, "JUNIO": 6,
    "JULIO": 7, "AGOSTO": 8, "SEPTIEMBRE": 9, "OCTUBRE": 10, "NOVIEMBRE": 11, "DICIEMBRE": 12,
}
ANIO = 2026

pat_grupo = re.compile(r"[A-ZÁÉÍÓÚ]+_G\d+")


def norm_nombre(s):
    if not s:
        return ""
    s = str(s).replace("\n", " ")
    return re.sub(r"\s+", " ", s).strip().upper()


def slugify(s):
    s = unicodedata.normalize("NFKD", s).encode("ascii", "ignore").decode()
    return re.sub(r"[^a-zA-Z0-9]+", "-", s).strip("-").lower()


def parse_fechas(raw):
    if not raw:
        return None, None
    t = str(raw).upper().replace("AL", " AL ").replace("-", " - ").replace("DE", " DE ")
    t = re.sub(r"\s+", " ", t)
    pares = re.findall(r"(\d{1,2})\s+(?:DE\s+)?(" + "|".join(MESES) + r")", t)
    fechas = []
    for dia, mes in pares:
        try:
            fechas.append(datetime.date(ANIO, MESES[mes], int(dia)).isoformat())
        except ValueError:
            pass
    if len(fechas) >= 2:
        return fechas[0], fechas[-1]
    if len(fechas) == 1:
        return fechas[0], None
    return None, None


def hora_iso(v):
    if isinstance(v, datetime.time):
        return v.strftime("%H:%M")
    if v in (None, "", "N/A"):
        return None
    return str(v).strip()


def headers(ws):
    raw = next(ws.iter_rows(min_row=1, max_row=1, values_only=True))
    return [(str(h).strip() if h and str(h).strip() else f"COL{i}") for i, h in enumerate(raw)]


def detectar_col_grupo(ws):
    conteo = defaultdict(int)
    for r in ws.iter_rows(min_row=2, max_row=200, values_only=True):
        for i, v in enumerate(r):
            if v and pat_grupo.search(str(v)):
                conteo[i] += 1
    return max(conteo, key=conteo.get) if conteo else None


def num_alumnos(v):
    """Acepta 90, 90.0, '8 ALUMNOS', etc. Devuelve int o None."""
    if v in (None, ""):
        return None
    if isinstance(v, (int, float)):
        return int(v)
    m = re.search(r"\d+", str(v))
    return int(m.group()) if m else None


def extraer_hoja(ws, plantel, sufijo, materias_set, grupos, prof_hist, alumnos_inline):
    hdr = headers(ws)
    cm = {h: i for i, h in enumerate(hdr)}
    idx_id = cm.get("ID")
    idx_doc = next((cm[k] for k in cm if "DOCENTE" in k.upper() and "FORMACIÓN" not in k.upper()), None)
    idx_mat = cm.get("MATERIA ") or cm.get("MATERIA")
    idx_plan = cm.get("PLAN DE ESTUDIOS")
    idx_cuatri = cm.get("CUATRIMESTRE")
    idx_tipo = cm.get("TIPO")
    idx_modal = cm.get("MODALIDAD")
    idx_dia = cm.get("DÍA")
    idx_turno = cm.get("TURNO")
    idx_hi = cm.get("Hora Inicio")
    idx_hf = cm.get("Hora Fin")
    idx_fec = cm.get("Fecha Automática")
    idx_conf = cm.get("Confirmación")
    idx_aula = cm.get("AULA ASIGNADA")
    idx_form = cm.get("FORMACIÓN DEL DOCENTE")
    idx_alum = next((i for h, i in cm.items() if h.upper() == "ALUMNOS"), None)
    idx_grupo = detectar_col_grupo(ws)

    slots = []
    seen = set()
    for r in ws.iter_rows(min_row=2, values_only=True):
        if idx_id is None or idx_id >= len(r) or r[idx_id] in (None, ""):
            continue
        sid = int(r[idx_id]) if isinstance(r[idx_id], (int, float)) else r[idx_id]
        materia = norm_nombre(r[idx_mat]) if idx_mat is not None else ""
        if not materia:
            continue  # fila sin materia = basura
        key = (sid, plantel, materia)
        if key in seen:
            continue
        seen.add(key)

        docente = norm_nombre(r[idx_doc]) if idx_doc is not None else ""
        es_coord = "COORD" in docente
        plan = norm_nombre(r[idx_plan]) if idx_plan is not None else ""
        grupo_raw = str(r[idx_grupo] or "").replace("\n", "").strip() if idx_grupo is not None and idx_grupo < len(r) else ""
        grupo = grupo_raw if pat_grupo.search(grupo_raw) else ""
        cuatri = str(r[idx_cuatri] or "").strip() if idx_cuatri is not None else ""
        tipo = str(r[idx_tipo] or "").strip().upper() if idx_tipo is not None else ""
        modal = str(r[idx_modal] or "").strip().upper() if idx_modal is not None else ""
        dia = str(r[idx_dia] or "").strip().upper() if idx_dia is not None else ""
        turno = str(r[idx_turno] or "").strip().upper() if idx_turno is not None else ""
        hi = hora_iso(r[idx_hi]) if idx_hi is not None and idx_hi < len(r) else None
        hf = hora_iso(r[idx_hf]) if idx_hf is not None and idx_hf < len(r) else None
        f_ini, f_fin = parse_fechas(r[idx_fec]) if idx_fec is not None and idx_fec < len(r) else (None, None)
        conf = str(r[idx_conf] or "").strip().upper() if idx_conf is not None and idx_conf < len(r) else ""
        aula = str(r[idx_aula] or "").strip() if idx_aula is not None and idx_aula < len(r) else ""
        formacion = norm_nombre(r[idx_form]) if idx_form is not None and idx_form < len(r) else ""
        alum = num_alumnos(r[idx_alum]) if idx_alum is not None and idx_alum < len(r) else None

        materias_set[materia] += 1
        if grupo:
            grupos.setdefault(grupo, {"clave": grupo, "plan": plan, "cuatrimestre": cuatri,
                                     "turno": turno, "plantel": plantel})
            if alum is not None:
                alumnos_inline[grupo] = alum

        if docente and not es_coord:
            prof_hist[docente]["materias"][materia] += 1
            prof_hist[docente]["planes"][plan] += 1
            if formacion:
                prof_hist[docente]["formaciones"][formacion] += 1
            prof_hist[docente]["slots"] += 1

        slots.append({
            "id_excel": sid, "plantel": plantel, "plan": plan,
            "cuatrimestre": cuatri, "grupo": grupo or None, "tipo": tipo,
            "materia": materia, "modalidad": modal,
            "docente": None if es_coord else (docente or None),
            "es_coordinador_virtual": es_coord,
            "formacion_docente": formacion or None, "aula": aula or None,
            "dia": dia or None, "turno": turno or None,
            "hora_inicio": hi, "hora_fin": hf,
            "fecha_inicio": f_ini, "fecha_fin": f_fin,
            "fecha_raw": str(r[idx_fec]).strip() if idx_fec is not None and idx_fec < len(r) and r[idx_fec] else None,
            "confirmacion": conf or None,
        })
    return slots, len([s for s in slots if s["grupo"]]), len([s for s in slots if not s["grupo"]])


def extraer_aulas(wb):
    def norm_tipo(t):
        t = (t or "").strip().lower()
        if t.startswith("teor"):
            return "Teoría"
        if t.startswith("pra") or t.startswith("prá"):
            return "Práctica"
        return t.title() if t else None

    def norm_clave(v):
        if v is None:
            return None
        s = str(v).strip()
        if re.fullmatch(r"\d+\.0", s):
            s = s[:-2]
        return s or None

    aulas, seen = [], set()
    if "Aulas" in wb.sheetnames:
        for r in wb["Aulas"].iter_rows(min_row=2, values_only=True):
            clave = norm_clave(r[0]) if r and len(r) > 0 else None
            if not clave or clave in seen:
                continue
            seen.add(clave)
            cap = None
            if len(r) > 2 and r[2] is not None:
                try:
                    cap = int(float(r[2]))
                except (ValueError, TypeError):
                    cap = None
            aulas.append({"clave": clave, "tipo": norm_tipo(r[1]) if len(r) > 1 else None, "capacidad": cap})
    return aulas


def main():
    wb = openpyxl.load_workbook(EXCEL, data_only=True)

    materias_set = Counter()
    grupos = {}
    prof_hist = defaultdict(lambda: {"materias": Counter(), "planes": Counter(),
                                     "formaciones": Counter(), "slots": 0})
    alumnos_inline = {}   # grupo -> alumnos (de la columna ALUMNOS de la propia hoja)
    todos_slots = []
    resumen_plantel = {}

    for hoja, (plantel, sufijo) in PLANTELES.items():
        if hoja not in wb.sheetnames:
            print(f"  !! hoja '{hoja}' no existe, se omite")
            continue
        slots, con_grupo, sin_grupo = extraer_hoja(
            wb[hoja], plantel, sufijo, materias_set, grupos, prof_hist, alumnos_inline)
        todos_slots.extend(slots)
        resumen_plantel[plantel] = {"slots": len(slots), "con_grupo": con_grupo, "sin_grupo": sin_grupo}

    aulas = extraer_aulas(wb)

    # ---- Alumnos por grupo: hoja ALUMNOS POR MATERIA (todos los planteles por clave) ----
    alumnos_por_grupo = dict(alumnos_inline)  # arranca con lo inline (OTAY/TECATE)
    if "ALUMNOS POR MATERIA" in wb.sheetnames:
        for r in wb["ALUMNOS POR MATERIA"].iter_rows(min_row=2, values_only=True):
            if not r or r[0] in (None, ""):
                continue
            clave = str(r[0]).replace("\n", "").strip()
            if clave not in grupos:   # ignora claves de otros sistemas / basura (ENS, ELIMINAR…)
                continue
            n = num_alumnos(r[1]) if len(r) > 1 else None
            if n is not None and clave not in alumnos_por_grupo:
                alumnos_por_grupo[clave] = n

    out = {
        "ciclo_historial": "2025-2026-3",
        "ciclo_a_asignar": "2026-2027-1",
        "resumen": {
            "slots_total": len(todos_slots),
            "por_plantel": resumen_plantel,
            "materias_distintas": len(materias_set),
            "grupos": len(grupos),
            "docentes_totales": len(prof_hist),
            "aulas": len(aulas),
            "grupos_con_alumnos": len(alumnos_por_grupo),
        },
        "planes_catalogo": sorted({g["plan"] for g in grupos.values() if g["plan"]} |
                                  {s["plan"] for s in todos_slots if s["plan"]}),
        "materias_catalogo": sorted(materias_set.keys()),
        "grupos": list(grupos.values()),
        "aulas": aulas,
        "alumnos_por_grupo": alumnos_por_grupo,
        "slots": todos_slots,
    }
    OUT.parent.mkdir(parents=True, exist_ok=True)
    OUT.write_text(json.dumps(out, ensure_ascii=False, indent=2))
    print(f"OK -> {OUT}\n")
    print(json.dumps(out["resumen"], indent=2, ensure_ascii=False))


if __name__ == "__main__":
    main()
