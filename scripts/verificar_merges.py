#!/usr/bin/env python3
"""Solo lectura. Pasada final: confirma que el forward-fill por celdas
combinadas reconstruye los registros de PALMAS y TECATE."""
import openpyxl, re
from collections import Counter
from pathlib import Path

EXCEL = Path.home() / "Downloads" / "PROYECCIÓN MAYO - AGOSTO -ACTUALIZADA.xlsx"
pat_grupo = re.compile(r"[A-ZÁÉÍÓÚ]+_G\d+")


def expandir_merges(ws):
    """Devuelve una matriz (lista de listas) con los valores de las celdas
    combinadas propagados a todas las celdas del rango."""
    filas = list(ws.iter_rows(values_only=True))
    matriz = [list(r) for r in filas]
    for rango in ws.merged_cells.ranges:
        min_r, min_c = rango.min_row, rango.min_col
        val = ws.cell(row=min_r, column=min_c).value
        for rr in range(rango.min_row, rango.max_row + 1):
            for cc in range(rango.min_col, rango.max_col + 1):
                ri, ci = rr - 1, cc - 1
                if 0 <= ri < len(matriz) and 0 <= ci < len(matriz[ri]):
                    matriz[ri][ci] = val
    return matriz


def analizar(nombre, wb):
    ws = wb[nombre]
    raw_hdr = matriz_hdr = next(ws.iter_rows(min_row=1, max_row=1, values_only=True))
    hdr = [(str(h).strip() if h and str(h).strip() else f"COL{i}") for i, h in enumerate(raw_hdr)]
    cm = {h: i for i, h in enumerate(hdr)}
    idx_id = cm.get("ID")
    idx_grupo = cm.get("GRUPO")
    idx_alum = [i for h, i in cm.items() if "ALUMN" in h.upper()]

    mat = expandir_merges(ws)
    n = id_ok = grp_ok = 0
    grupos = set()
    alum_fmt = Counter()
    ejemplos = []
    for r in mat[1:]:
        if not any(v not in (None, "") for v in r):
            continue
        n += 1
        if idx_id is not None and idx_id < len(r) and r[idx_id] not in (None, ""):
            id_ok += 1
        if idx_grupo is not None and idx_grupo < len(r) and r[idx_grupo] not in (None, ""):
            g = str(r[idx_grupo]).replace("\n", "").strip()
            if pat_grupo.search(g):
                grp_ok += 1
                grupos.add(g)
        for ia in idx_alum:
            if ia < len(r) and r[ia] not in (None, ""):
                v = r[ia]
                if isinstance(v, (int, float)):
                    alum_fmt["numero"] += 1
                elif re.search(r"\d", str(v)):
                    alum_fmt["texto:" + re.sub(r"\d+", "N", str(v).strip())[:18]] += 1
                else:
                    alum_fmt["texto_sin_num"] += 1
        if len(ejemplos) < 4 and idx_grupo is not None and idx_grupo < len(r) and r[idx_grupo]:
            ejemplos.append((r[idx_id] if idx_id is not None else None,
                             str(r[idx_grupo]).replace("\n", " ").strip(),
                             r[cm.get("MATERIA", 0)] if "MATERIA" in cm else None))
    print(f"\n--- {nombre} (con merges expandidos) ---")
    print(f"  filas={n}  ID lleno={id_ok} ({100*id_ok//max(n,1)}%)"
          f"  GRUPO válido={grp_ok} ({100*grp_ok//max(n,1)}%)  grupos distintos={len(grupos)}")
    print(f"  formatos ALUMNOS: {dict(alum_fmt)}")
    print("  ejemplos (id, grupo, materia):")
    for e in ejemplos:
        print(f"     {e}")


def main():
    wb = openpyxl.load_workbook(EXCEL, data_only=True)
    print("=" * 70)
    print("PASADA FINAL: forward-fill de celdas combinadas (PALMAS, TECATE)")
    print("  Antes (sin expandir): PALMAS GRUPO=10%, TECATE GRUPO=66%")
    print("=" * 70)
    for nombre in ["PALMAS", "TECATE", "OTAY"]:
        analizar(nombre, wb)


if __name__ == "__main__":
    main()
