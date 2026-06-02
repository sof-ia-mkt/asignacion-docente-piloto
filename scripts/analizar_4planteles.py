#!/usr/bin/env python3
"""Solo lectura. Verifica los datos de los 4 planteles antes de migrar.
No escribe nada. Pasadas 4 y 5 de la revisión exigida."""
import openpyxl, re
from collections import Counter, defaultdict
from pathlib import Path

EXCEL = Path.home() / "Downloads" / "PROYECCIÓN MAYO - AGOSTO -ACTUALIZADA.xlsx"
HOJAS_PLANTEL = ["CASA BLANCA", "PALMAS", "OTAY", "TECATE"]
pat_grupo = re.compile(r"[A-ZÁÉÍÓÚ]+_G\d+")


def hdr_de(ws):
    raw = next(ws.iter_rows(min_row=1, max_row=1, values_only=True))
    return [(str(h).strip() if h and str(h).strip() else f"COL{i}") for i, h in enumerate(raw)]


def main():
    wb = openpyxl.load_workbook(EXCEL, data_only=True)
    print("=" * 70)
    print("PASADA 4: estructura por plantel (headers + columnas clave)")
    print("=" * 70)
    for nombre in HOJAS_PLANTEL:
        ws = wb[nombre]
        hdr = hdr_de(ws)
        # detectar columna grupo por patrón
        conteo = defaultdict(int)
        filas = 0
        for r in ws.iter_rows(min_row=2, values_only=True):
            if not any(v not in (None, "") for v in r):
                continue
            filas += 1
            for i, v in enumerate(r):
                if v and pat_grupo.search(str(v)):
                    conteo[i] += 1
        idx_grupo = max(conteo, key=conteo.get) if conteo else None
        print(f"\n--- {nombre}  (filas con datos={filas}) ---")
        print(f"  headers: {hdr}")
        print(f"  col grupo detectada: idx={idx_grupo}"
              f" (header='{hdr[idx_grupo] if idx_grupo is not None else '?'}',"
              f" celdas con patrón={conteo.get(idx_grupo,0)})")

    print("\n" + "=" * 70)
    print("PASADA 4b: celdas combinadas (merged) por hoja de plantel")
    print("=" * 70)
    for nombre in HOJAS_PLANTEL:
        ws = wb[nombre]
        merges = ws.merged_cells.ranges
        print(f"  {nombre}: {len(merges)} rangos combinados")
        for m in list(merges)[:5]:
            print(f"      {m}")

    print("\n" + "=" * 70)
    print("PASADA 5: ID / GRUPO / ALUMNOS — completitud y formatos por plantel")
    print("=" * 70)
    for nombre in HOJAS_PLANTEL:
        ws = wb[nombre]
        hdr = hdr_de(ws)
        cm = {h: i for i, h in enumerate(hdr)}
        idx_id = cm.get("ID")
        # col grupo
        conteo = defaultdict(int)
        for r in ws.iter_rows(min_row=2, max_row=200, values_only=True):
            for i, v in enumerate(r):
                if v and pat_grupo.search(str(v)):
                    conteo[i] += 1
        idx_grupo = max(conteo, key=conteo.get) if conteo else None
        # cols que mencionan ALUMNO
        idx_alumnos = [i for h, i in cm.items() if "ALUMN" in h.upper()]

        n = id_ok = grp_ok = 0
        grupos = set()
        alumnos_fmt = Counter()
        for r in ws.iter_rows(min_row=2, values_only=True):
            if not any(v not in (None, "") for v in r):
                continue
            n += 1
            if idx_id is not None and r[idx_id] not in (None, ""):
                id_ok += 1
            if idx_grupo is not None and r[idx_grupo] not in (None, ""):
                grp_ok += 1
                g = str(r[idx_grupo]).replace("\n", "").strip()
                if pat_grupo.search(g):
                    grupos.add(g)
            for ia in idx_alumnos:
                v = r[ia] if ia < len(r) else None
                if v not in (None, ""):
                    if isinstance(v, (int, float)):
                        alumnos_fmt["numero"] += 1
                    elif re.search(r"\d", str(v)):
                        alumnos_fmt["texto_con_digito"] += 1
                    else:
                        alumnos_fmt["texto_sin_digito"] += 1
        print(f"\n--- {nombre} ---")
        print(f"  filas={n}  ID lleno={id_ok} ({100*id_ok//max(n,1)}%)"
              f"  GRUPO lleno={grp_ok} ({100*grp_ok//max(n,1)}%)  grupos distintos={len(grupos)}")
        print(f"  cols ALUMNOS: {[hdr[i] for i in idx_alumnos]}  formatos={dict(alumnos_fmt)}")

    print("\n" + "=" * 70)
    print("PASADA 5b: ALUMNOS POR MATERIA — sufijos de plantel")
    print("=" * 70)
    if "ALUMNOS POR MATERIA" in wb.sheetnames:
        ws = wb["ALUMNOS POR MATERIA"]
        suf = Counter()
        total = 0
        for r in ws.iter_rows(min_row=2, values_only=True):
            if not r or r[0] in (None, ""):
                continue
            total += 1
            g = str(r[0]).replace("\n", "").strip()
            m = re.search(r"_([A-Za-z.]+)$", g)
            suf[m.group(1) if m else "(sin sufijo)"] += 1
        print(f"  filas={total}")
        for s, c in suf.most_common():
            print(f"    {s:14s} {c}")

    print("\n" + "=" * 70)
    print("PASADA 5c: Materias (catálogo cross-campus) — completitud")
    print("=" * 70)
    if "Materias" in wb.sheetnames:
        ws = wb["Materias"]
        hdr = hdr_de(ws)
        print(f"  headers: {hdr}")
        n = 0
        lleno = Counter()
        for r in ws.iter_rows(min_row=2, values_only=True):
            if not any(v not in (None, "") for v in r):
                continue
            n += 1
            for i, v in enumerate(r):
                if v not in (None, ""):
                    lleno[hdr[i]] += 1
        print(f"  filas={n}")
        for h in hdr:
            print(f"    {h:30s} {lleno.get(h,0)} ({100*lleno.get(h,0)//max(n,1)}%)")


if __name__ == "__main__":
    main()
