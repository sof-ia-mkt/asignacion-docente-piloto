#!/usr/bin/env python3
"""
Extrae y limpia los datos de CASA BLANCA desde el Excel de proyección.
Salida: db/seed_data/casablanca.json (consumido por el seed a Supabase).

Decisiones de limpieza:
- Nombres de docente: quitar \n, colapsar espacios, upper.
- Llave de slot: ID + plantel (el ID NO es único entre planteles).
- Fechas: parser tolerante para ~29 formatos en texto libre.
- Grupo: columna sin header (idx detectado por patrón PLAN_Gnn).
- Mayo = historial. Septiembre = misma estructura, docente vacío.
"""
import openpyxl, re, json, datetime, unicodedata
from collections import Counter, defaultdict
from pathlib import Path

EXCEL = Path.home() / "Downloads" / "PROYECCIÓN MAYO - AGOSTO -ACTUALIZADA.xlsx"
OUT = Path(__file__).resolve().parent.parent / "db" / "seed_data" / "casablanca.json"
N_DOCENTES_PILOTO = 20

MESES = {
    "ENERO": 1, "FEBRERO": 2, "MARZO": 3, "ABRIL": 4, "MAYO": 5, "JUNIO": 6,
    "JULIO": 7, "AGOSTO": 8, "SEPTIEMBRE": 9, "OCTUBRE": 10, "NOVIEMBRE": 11, "DICIEMBRE": 12,
}
ANIO = 2026  # ciclo 2025-2026-3 corre may-ago 2026


def norm_nombre(s):
    if not s:
        return ""
    s = str(s).replace("\n", " ")
    s = re.sub(r"\s+", " ", s).strip().upper()
    return s


def slugify(s):
    s = unicodedata.normalize("NFKD", s).encode("ascii", "ignore").decode()
    s = re.sub(r"[^a-zA-Z0-9]+", "-", s).strip("-").lower()
    return s


def parse_fechas(raw):
    """Devuelve (fecha_inicio, fecha_fin) ISO o (None, None)."""
    if not raw:
        return None, None
    t = str(raw).upper()
    t = t.replace("AL", " AL ").replace("-", " - ").replace("DE", " DE ")
    t = re.sub(r"\s+", " ", t)
    # encuentra pares (dia, mes)
    pares = re.findall(r"(\d{1,2})\s+(?:DE\s+)?(" + "|".join(MESES) + r")", t)
    fechas = []
    for dia, mes in pares:
        try:
            fechas.append(datetime.date(ANIO, MESES[mes], int(dia)).isoformat())
        except ValueError:
            pass
    if len(fechas) >= 2:
        return fechas[0], fechas[-1]
    if len(fechas) == 1:
        return fechas[0], None
    return None, None


def hora_iso(v):
    if isinstance(v, datetime.time):
        return v.strftime("%H:%M")
    if v in (None, "", "N/A"):
        return None
    return str(v).strip()


def main():
    wb = openpyxl.load_workbook(EXCEL, data_only=True)
    ws = wb["CASA BLANCA"]
    raw_hdr = next(ws.iter_rows(min_row=1, max_row=1, values_only=True))
    hdr = [(str(h).strip() if h and str(h).strip() else f"COL{i}") for i, h in enumerate(raw_hdr)]
    cm = {h: i for i, h in enumerate(hdr)}

    idx_id = cm["ID"]
    idx_doc = [cm[k] for k in cm if "DOCENTE" in k][0]
    idx_mat = cm.get("MATERIA ") or cm.get("MATERIA")
    idx_plan = cm["PLAN DE ESTUDIOS"]
    idx_cuatri = cm["CUATRIMESTRE"]
    idx_tipo = cm["TIPO"]
    idx_modal = cm.get("MODALIDAD")
    idx_dia = cm.get("DÍA")
    idx_turno = cm.get("TURNO")
    idx_hi = cm.get("Hora Inicio")
    idx_hf = cm.get("Hora Fin")
    idx_fec = cm.get("Fecha Automática")
    idx_conf = cm.get("Confirmación")
    idx_aula = cm.get("AULA ASIGNADA")
    idx_form = cm.get("FORMACIÓN DEL DOCENTE")

    # detectar columna de grupo por patrón
    pat_grupo = re.compile(r"[A-ZÁÉÍÓÚ]+_G\d+")
    conteo = defaultdict(int)
    for r in ws.iter_rows(min_row=2, max_row=80, values_only=True):
        for i, v in enumerate(r):
            if v and pat_grupo.search(str(v)):
                conteo[i] += 1
    idx_grupo = max(conteo, key=conteo.get)

    slots = []
    seen = set()
    prof_hist = defaultdict(lambda: {"materias": Counter(), "planes": Counter(),
                                      "formaciones": Counter(), "slots": 0})
    materias_set = Counter()
    grupos = {}

    for r in ws.iter_rows(min_row=2, values_only=True):
        if not r[idx_id]:
            continue
        sid = int(r[idx_id]) if isinstance(r[idx_id], (int, float)) else r[idx_id]
        key = (sid, "CASA BLANCA")
        if key in seen:
            continue
        seen.add(key)

        docente = norm_nombre(r[idx_doc])
        es_coord = "COORD" in docente
        materia = norm_nombre(r[idx_mat])
        plan = norm_nombre(r[idx_plan])
        grupo = str(r[idx_grupo] or "").replace("\n", "").strip() if idx_grupo is not None else ""
        cuatri = str(r[idx_cuatri] or "").strip()
        tipo = str(r[idx_tipo] or "").strip().upper()
        modal = str(r[idx_modal] or "").strip().upper() if idx_modal is not None else ""
        dia = str(r[idx_dia] or "").strip().upper() if idx_dia is not None else ""
        turno = str(r[idx_turno] or "").strip().upper() if idx_turno is not None else ""
        hi = hora_iso(r[idx_hi]) if idx_hi is not None else None
        hf = hora_iso(r[idx_hf]) if idx_hf is not None else None
        f_ini, f_fin = parse_fechas(r[idx_fec]) if idx_fec is not None else (None, None)
        conf = str(r[idx_conf] or "").strip().upper() if idx_conf is not None else ""
        aula = str(r[idx_aula] or "").strip() if idx_aula is not None else ""
        formacion = norm_nombre(r[idx_form]) if idx_form is not None else ""

        materias_set[materia] += 1
        if grupo and pat_grupo.search(grupo):
            grupos.setdefault(grupo, {"clave": grupo, "plan": plan, "cuatrimestre": cuatri, "turno": turno})

        if docente and not es_coord:
            prof_hist[docente]["materias"][materia] += 1
            prof_hist[docente]["planes"][plan] += 1
            if formacion:
                prof_hist[docente]["formaciones"][formacion] += 1
            prof_hist[docente]["slots"] += 1

        slots.append({
            "id_excel": sid,
            "plantel": "CASA BLANCA",
            "plan": plan,
            "cuatrimestre": cuatri,
            "grupo": grupo,
            "tipo": tipo,
            "materia": materia,
            "modalidad": modal,
            "docente": None if es_coord else (docente or None),
            "es_coordinador_virtual": es_coord,
            "formacion_docente": formacion or None,
            "aula": aula or None,
            "dia": dia or None,
            "turno": turno or None,
            "hora_inicio": hi,
            "hora_fin": hf,
            "fecha_inicio": f_ini,
            "fecha_fin": f_fin,
            "fecha_raw": str(r[idx_fec]).strip() if idx_fec is not None and r[idx_fec] else None,
            "confirmacion": conf or None,
        })

    # ---- Aulas (catálogo de salones) ----
    def norm_tipo_aula(t):
        t = (t or "").strip().lower()
        if t.startswith("teor"):
            return "Teoría"
        if t.startswith("pra") or t.startswith("prá"):
            return "Práctica"
        return t.title() if t else None

    def norm_clave_aula(v):
        if v is None:
            return None
        s = str(v).strip()
        if re.fullmatch(r"\d+\.0", s):  # "104.0" -> "104"
            s = s[:-2]
        return s or None

    aulas = []
    seen_aula = set()
    if "Aulas" in wb.sheetnames:
        for r in wb["Aulas"].iter_rows(min_row=2, values_only=True):
            clave = norm_clave_aula(r[0]) if r and len(r) > 0 else None
            if not clave or clave in seen_aula:
                continue
            seen_aula.add(clave)
            cap = None
            if len(r) > 2 and r[2] is not None:
                try:
                    cap = int(float(r[2]))
                except (ValueError, TypeError):
                    cap = None
            aulas.append({
                "clave": clave,
                "tipo": norm_tipo_aula(r[1]) if len(r) > 1 else None,
                "capacidad": cap,
            })

    # ---- Alumnos por grupo (solo grupos de CASA BLANCA) ----
    alumnos_por_grupo = {}
    if "ALUMNOS POR MATERIA" in wb.sheetnames:
        for r in wb["ALUMNOS POR MATERIA"].iter_rows(min_row=2, values_only=True):
            if not r or r[0] is None:
                continue
            clave = str(r[0]).replace("\n", "").strip()
            if clave not in grupos:  # ignora grupos de otros planteles
                continue
            if len(r) > 1 and r[1] is not None:
                try:
                    alumnos_por_grupo[clave] = int(float(r[1]))
                except (ValueError, TypeError):
                    pass

    # top 20 docentes por slots, luego por materias distintas
    rank = sorted(prof_hist.items(), key=lambda kv: (-kv[1]["slots"], -len(kv[1]["materias"])))
    piloto = []
    for nombre, info in rank[:N_DOCENTES_PILOTO]:
        piloto.append({
            "nombre": nombre,
            "slug": slugify(nombre),
            "slots_mayo": info["slots"],
            "historial_materias": [m for m, _ in info["materias"].most_common()],
            "planes": [p for p, _ in info["planes"].most_common()],
            "formacion_excel": info["formaciones"].most_common(1)[0][0] if info["formaciones"] else None,
        })

    out = {
        "plantel": "CASA BLANCA",
        "ciclo_historial": "2025-2026-3",
        "ciclo_a_asignar": "2026-2027-1",
        "resumen": {
            "slots": len(slots),
            "materias_distintas": len(materias_set),
            "grupos": len(grupos),
            "docentes_totales": len(prof_hist),
            "docentes_piloto": len(piloto),
            "aulas": len(aulas),
            "grupos_con_alumnos": len(alumnos_por_grupo),
        },
        "materias_catalogo": sorted(materias_set.keys()),
        "grupos": list(grupos.values()),
        "aulas": aulas,
        "alumnos_por_grupo": alumnos_por_grupo,
        "slots_mayo": slots,
        "docentes_piloto": piloto,
    }
    OUT.parent.mkdir(parents=True, exist_ok=True)
    OUT.write_text(json.dumps(out, ensure_ascii=False, indent=2))
    print(f"OK -> {OUT}")
    print(json.dumps(out["resumen"], indent=2, ensure_ascii=False))
    print("\nDocentes piloto:")
    for p in piloto:
        print(f"  {p['nombre'][:38]:38s} mayo={p['slots_mayo']:2d}  hist={len(p['historial_materias'])} materias")


if __name__ == "__main__":
    main()
